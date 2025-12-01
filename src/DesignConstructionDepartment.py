import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from psycopg2 import sql
from minio import Minio
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from psycopg2.extras import RealDictCursor
import io
import openpyxl
import re

def DCD_WarehouseInstallation_Upload_function(conn, owner, file, minio_client, MINIO_BUCKET):
    try:
        project_code = file.filename.split(".")[0]
        print(project_code)
        wb = openpyxl.load_workbook(file.file)
        ws = wb.active

        # --- Lấy header ---
        headers = {cell.value: cell.column for cell in ws[1]}

        # Các cột khóa chính
        col_higher = headers.get("Higher-level function with preceding sign")
        col_mount = headers.get("Mounting location with preceding sign")
        col_dt = headers.get("DT (identifying, without project structures, with preceding sign)")
        col_main_function = headers.get(f"Main function")

        if not all([col_higher, col_mount, col_dt]):
            raise ValueError("Thiếu một trong các cột khóa chính (Higher-level / Mounting / DT identifying).")

        # Tập hợp các cột Order và Item cần xử lý
        col_pairs = []
        for i in range(1, 7):
            order_col = headers.get(f"Order number [{i}]")
            item_col = headers.get(f"Item number [{i}]")
            if order_col and item_col :
                col_pairs.append((order_col, item_col))

        if not col_pairs:
            raise ValueError("Không tìm thấy các cột Order number [n] và Item number [n] trong file Excel!")

        # --- Duyệt từng dòng ---
        for row in range(2, ws.max_row + 1):
            higher = ws.cell(row=row, column=col_higher).value
            mount = ws.cell(row=row, column=col_mount).value
            dt_ident = ws.cell(row=row, column=col_dt).value
            main_func_value = ws.cell(row=row, column=col_main_function).value if col_main_function else None

            if not (higher and mount and dt_ident):
                continue
            if main_func_value != "1":
                continue

            for order_col, item_col in col_pairs:
                order_num = ws.cell(row=row, column=order_col).value
                if not order_num:
                    continue

                with conn.cursor() as cur:
                    # if higher == "=LX931EA00.E06":
                    #     print(f"higher: {higher}, mount: {mount}, dt_ident: {dt_ident}, order_num: {order_num}")
                    cur.execute("""
                        SELECT "seri_number"
                        FROM "WS_Installation"
                        WHERE "higher_lever_function" = %s
                            AND "location" = %s
                            AND "dt" = %s
                            AND "project_code" = %s
                            AND "part_no" = %s
                        LIMIT 1
                    """, (higher, mount, dt_ident, project_code, order_num))
                    result = cur.fetchone()
                    if result is not None:
                        print(result)

                seri = result[0] if result else None
                ws.cell(row=row, column=item_col, value=seri)

        # --- Ghi kết quả ra memory ---
        output_path = f"./minio/minio_data/DCD/Output/{project_code}.xlsx"
        wb.save(output_path)

        object_name = f"data/DCD/Installation/Output/{project_code}.xlsx"
        minio_client.fput_object(
            bucket_name=MINIO_BUCKET,
            object_name=object_name,
            file_path=output_path
        )

        presigned_url = minio_client.presigned_get_object(
            bucket_name=MINIO_BUCKET,
            object_name=object_name,
            expires=timedelta(seconds=3600)
        )

        return {
            "status": "success",
            "url": presigned_url
        }

    except Exception as e:
        raise RuntimeError(f"Lỗi xử lý file Excel: {e}")
