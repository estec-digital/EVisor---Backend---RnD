import pandas as pd
from io import BytesIO
from minio import Minio
from pydantic import BaseModel
from typing import List, Optional
import numpy as np
import openpyxl
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Border, Side
from fastapi.responses import JSONResponse
from openpyxl.styles import Alignment
import os
import jpype, sys
import os, re, json
import tempfile
from collections import Counter

# def issummary_file(content: bytes) -> bool:
#     df = pd.read_excel(BytesIO(content))
#     # Các cột mong muốn
#     expected_columns = [
#         "STT",
#         "Tên nhân sự",
#         "Mã dự án",
#         "Mô tả công việc",
#         "Tên nhân sự - filter",
#         "Mã dự án - filter",
#         "Thời gian bắt đầu",
#         "Thời gian kết thúc"
#     ]
#     # Chuyển df.columns thành set để kiểm tra dễ hơn
#     actual_columns = set(df.columns)
#     # Kiểm tra tất cả các cột mong muốn đều có trong df
#     if all(col in actual_columns for col in expected_columns):
#         return True  # Là file summary
#     else:
#         return False  # Không phải file summary
def duplicate_project_code(list_codes: List[List[str]]) -> bool:
    print("list_codes:", list_codes)
    counter = Counter(list_codes)
    duplicates = [code for code, count in counter.items() if count > 1]
    print("duplicates:", duplicates)
    return duplicates

def issummary_file(content: bytes, filename: str, jvm_path) -> bool:
    ext = os.path.splitext(filename)[1].lower()
    
    if ext in [".xlsx", ".xls", ".csv"]:
        # --- Xử lý Excel ---
        df = pd.read_excel(BytesIO(content))
        actual_columns = set(df.columns)
        expected_columns = [
            "STT",
            "Tên nhân sự",
            "Mã dự án",
            "Mô tả công việc",
            "Tên nhân sự - filter",
            "Mã dự án - filter",
            "Thời gian bắt đầu",
            "Thời gian kết thúc"
        ]
        if all(col in actual_columns for col in expected_columns):
            codes = df["Mã dự án"].dropna().unique().tolist()
        else:
            # print(df.head())
            codes = [df.iloc[4, 4]]
        return all(col in actual_columns for col in expected_columns), codes

    elif ext == ".mpp":
        # --- Lưu file tạm và tự xóa---
        with tempfile.NamedTemporaryFile(delete=False, suffix=".mpp") as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            actual_columns, project_code = mpp_isissummary_file(
                mpp_file=tmp_path,
                mpxj_dir="./mpxj",
                json_output_path=None,
                jvm_path=jvm_path
            )
            print("Parsed JSON from MPP:", actual_columns)
            expected_columns = [
                "Hours", 
                "Unit", 
                "Name", 
                "Resource Names", 
                "Notes", 
                "Finish", 
                "Working place", 
                "Start", 
                "Duration", 
                "ID", 
                "% Complete"
            ]
            project_code = [project_code] if project_code else []
            # print("Extracted project codes:", codes)
            return all(col not in actual_columns for col in expected_columns), project_code
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
    else:
        return False

def mpp_isissummary_file(mpp_file, mpxj_dir, json_output_path, jvm_path):
    # --- Build classpath ---
    lib_path = os.path.join(mpxj_dir, "lib")
    mpxj_jar = os.path.join(mpxj_dir, "mpxj.jar")
    jars = [os.path.join(lib_path, jar) for jar in os.listdir(lib_path) if jar.endswith(".jar")]
    jars.append(mpxj_jar)
    classpath = ";".join(jars)

    # --- Start JVM nếu chưa chạy ---
    if not jpype.isJVMStarted():
        jpype.startJVM(jvm_path, "-ea", f"-Djava.class.path={classpath}", convertStrings=True)

    # --- Đọc MPP ---
    MPPReader = jpype.JClass("org.mpxj.mpp.MPPReader")
    reader = MPPReader()
    project = reader.read(mpp_file)

    # === Task Data Extraction
    task_data = []
    project_code = None
    for task in project.getTasks():
        if task.getName() is None:
            continue
        if task.getID() == 0 and project_code is None:
            project_code = extract_project_code(task.getName())

        task_data.append({
        "ID": task.getID(),
        "Name": task.getName(),
        "Start": str(task.getStart()).split("T")[0] if task.getStart() else "",
        "Finish": str(task.getFinish()).split("T")[0] if task.getFinish() else "",
        "Duration": str(task.getDuration()),
        "% Complete": task.getPercentageComplete(),
        "Resource Names": ", ".join([
            r.getName()
            for a in task.getResourceAssignments()
            if (r := a.getResource()) and r.getName()
        ]) or "unassigned",
        "Working place": task.getText(3),
        "Hours": task.getText(2),
        "Unit": task.getText(4),
        "Notes": task.getNotes()
        })

    if not task_data:
        return []

    df = pd.DataFrame(task_data)
    # print("project_code:", project_code)
    return set(df.columns), project_code

def POD_TimeTracker_Merge_Manual_function(minio_client: Minio, input: BaseModel, jvm_path = "C:/Program Files/Eclipse Adoptium/jdk-21.0.7.6-hotspot/bin/server/jvm.dll"):
    try:
        input_dict = input.dict()
        request_id = input_dict.get("request_id", "evisor-1234567890")
        user_id = input_dict.get("user_id", "evisor")
        start_time = input_dict.get("start_time", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        summary_file = input_dict.get("summary_file", None).split("estec/")[-1]
        print(f"Summary file: {summary_file}")
        path_files = input_dict.get("path_files", [])
        duplicate = input_dict.get("duplicate", [])

        obj = minio_client.get_object("estec", summary_file)
        data = obj.read()
        file_stream = BytesIO(data)

        df_final = pd.read_excel(file_stream, engine='openpyxl')
        cols_to_fill = [
                "STT",
                "Tên nhân sự",
                "Mã dự án",
                "Mô tả công việc",
                "Tên nhân sự - filter",
                "Mã dự án - filter",
                "Thời gian bắt đầu",
                "Thời gian kết thúc"
            ]
        df_final[cols_to_fill] = df_final[cols_to_fill].fillna(method="ffill")
        results = []
        print(f"Summary file: {summary_file}")

        if "Mã dự án - filter" in df_final.columns:
            if duplicate:  # Nếu list duplicate không rỗng
                before = len(df_final)
                df_final = df_final[~df_final["Mã dự án - filter"].isin(duplicate)].reset_index(drop=True)
                after = len(df_final)
                print(f"🔎 Đã loại bỏ {before - after} dòng có Mã dự án - filter nằm trong duplicate")

        for file_path in path_files:
            obj = minio_client.get_object("estec", file_path)
            data = obj.read()
            file_stream = BytesIO(data)
            # json = processing_json(file_stream)

            # Phân loại theo định dạng file
            if file_path.lower().endswith((".xls", ".xlsx")):
                json = processing_json(file_stream)

            elif file_path.lower().endswith(".mpp"):
                # Lưu tạm file .mpp xuống local để MPXJ đọc
                tmp_path = f"{os.path.basename(file_path)}"
                with open(tmp_path, "wb") as f:
                    f.write(data)
                # Gọi hàm đọc MPP (bạn đã viết)
                mpxj_dir = "./mpxj"  # thư mục chứa mpxj.jar + lib/
                json = mpp_file_processing(tmp_path, mpxj_dir, None, jvm_path)
                
            if isinstance(json, dict) and json.get("status") == "error":
                return json
            results.append(json)

            stt_max = df_final["STT"].max() if not df_final.empty else 0
            df = generate_dataframe(json, stt_max)
            print("df_final:", df_final.head())
            print(f"DataFrame: {df}")
            df_final = pd.concat([df_final, df], ignore_index=True)

        df_final = df_final.sort_values(by=["Tên nhân sự", "Mã dự án", "Thời gian bắt đầu"]).reset_index(drop=True)
        print(f"DataFrame : {df}")

        df_final = df_final.drop_duplicates(subset=["Tên nhân sự", "Mã dự án", "Mô tả công việc", "Thời gian bắt đầu", "Thời gian kết thúc"]).reset_index(drop=True)
        print("df_final:", df_final.head())
        output_path_local, overwork = save_file_local(df_final)
        output_path_minio = save_file_minio(minio_client, output_path_local)
        # print(f"Output path: {output_path_minio}")

        return {
            "status": "success",
            "request_id": request_id,
            "user_id": user_id,
            "start_time": start_time,
            "output": output_path_minio,
            "overwork": overwork if overwork else None      
        }

    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

def POD_TimeTracker_Download_function(minio_client: Minio, input: BaseModel, MINIO_BUCKET: str):
    try:
        path_file = input.path_file
        path_file = path_file.split(f"{MINIO_BUCKET}/")[-1] if f"{MINIO_BUCKET}/" in path_file else path_file
        url = minio_client.presigned_get_object(MINIO_BUCKET, path_file, expires=timedelta(seconds=3600))
        return {
            "status": "success",
            "url": url
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

def POD_TimeTracker_Getfile_function(minio_client: Minio, input: BaseModel, MINIO_BUCKET: str):
    """
        Hàm để lấy file từ MinIO và chuyển đổi nội dung sang định dạng JSON.
        Tham số:
        - minio_client: Đối tượng Minio để tương tác với MinIO.
        - input: Đối tượng chứa thông tin đầu vào, bao gồm request_id, user_id và path_file.
        Trả về:
        - dict: Kết quả xử lý, bao gồm status và dữ liệu JSON.
        Nếu có lỗi xảy ra, trả về dict chứa status là "error" và message mô tả lỗi.
    """
    try:
        path_file = input.path_file
        path_file = path_file.split(f"{MINIO_BUCKET}/")[-1] if f"{MINIO_BUCKET}/" in path_file else path_file
        # print(f"Retrieving file from MinIO: {path_file}")
        response = minio_client.get_object(MINIO_BUCKET, path_file)
        try:
            content = BytesIO(response.read())
            df = pd.read_excel(content)
            df.replace({np.nan: None, np.inf: None, -np.inf: None}, inplace=True)
            json_data = df.to_dict(orient="records")
            # print(f"Retrieved data from MinIO: {json_data}")
            return JSONResponse(content={"status": "success", "data": json_data})
        finally:
            response.close()
            response.release_conn()
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

def POD_TimeTracker_Merge_function(minio_client: Minio, input: BaseModel, jvm_path = "C:/Program Files/Eclipse Adoptium/jdk-21.0.7.6-hotspot/bin/server/jvm.dll"):
    """
    Hàm xử lý dữ liệu từ MinIO và trả về kết quả.
    Tham số:
    - minio_client: Đối tượng Minio để tương tác với MinIO.
    - input: Đối tượng chứa thông tin đầu vào, bao gồm request_id, user_id, start_time và path_files.
    Trả về:
    - dict: Kết quả xử lý, bao gồm status, request_id, user_id, start_time, output và overwork (nếu có).
    Nếu có lỗi xảy ra, trả về dict chứa status là "error" và message mô tả lỗi.
    """
    try:
        input_dict = input.dict()
        request_id = input_dict.get("request_id", "evisor-1234567890")
        user_id = input_dict.get("user_id", "evisor")
        start_time = input_dict.get("start_time")
        path_files = input_dict.get("path_files", [])
        duplicate = input_dict.get("duplicate", [])
        summary_file = input_dict.get("summary_file", None)

        if not path_files:
            return {"status": "error", "message": "No files provided in path_files."}
        
        df_final = pd.DataFrame()
        results = []

        print(f"Summary file: {summary_file}")
        if summary_file:
            df_final = pd.read_excel(BytesIO(minio_client.get_object("estec", summary_file).read()), engine='openpyxl')
            print("TEST",df_final.head())

            df_final = df_final[~df_final["Mã dự án - filter"].isin(duplicate)].reset_index(drop=True)

        for file_path in path_files:
            obj = minio_client.get_object("estec", file_path)
            data = obj.read()
            file_stream = BytesIO(data)

            # Phân loại theo định dạng file
            if file_path.lower().endswith((".xls", ".xlsx")):
                json_data = processing_json(file_stream)

            elif file_path.lower().endswith(".mpp"):
                # Lưu tạm file .mpp xuống local để MPXJ đọc
                tmp_path = f"{os.path.basename(file_path)}"
                with open(tmp_path, "wb") as f:
                    f.write(data)
                # Gọi hàm đọc MPP (bạn đã viết)
                mpxj_dir = "./mpxj"  # thư mục chứa mpxj.jar + lib/
                json_data = mpp_file_processing(tmp_path, mpxj_dir, None, jvm_path)
            else:
                return {"status": "error", "message": f"Unsupported file type: {file_path}"}
            
            if isinstance(json_data, dict) and json_data.get("status") == "error":
                return json_data
            
            results.append(json_data)
            stt_max = df_final["STT"].max() if not df_final.empty else 0
            df = generate_dataframe(json_data, stt_max)
            df_final = pd.concat([df_final, df], ignore_index=True) 
        
        # df_final = df_final.sort_values(by=["STT", "Tên nhân sự", "Mã dự án", "Thời gian bắt đầu"]).reset_index(drop=True)
        df_final = df_final.sort_values(by=["Tên nhân sự", "Mã dự án", "Thời gian bắt đầu"]).reset_index(drop=True)
        df_final = df_final.drop_duplicates(subset=["Tên nhân sự", "Mã dự án", "Mô tả công việc", "Thời gian bắt đầu", "Thời gian kết thúc"]).reset_index(drop=True)
        print("df_final:", df_final.head())
        output_path_local, overwork = save_file_local(df_final)
        output_path_minio = save_file_minio(minio_client, output_path_local)
        # print(f"Output path: {output_path_minio}")

        return {
            "status": "success",
            "request_id": request_id,
            "user_id": user_id,
            "start_time": start_time,
            "output": output_path_minio,
            "overwork": overwork if overwork else None      
        }

    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

def mpp_file_processing(mpp_file, mpxj_dir,json_output_path,jvm_path):
    # === CONFIGURATION ===
    lib_path = os.path.join(mpxj_dir, "lib")
    mpxj_jar = os.path.join(mpxj_dir, "mpxj.jar")
    log_config_path = os.path.abspath("./mpxj/log4j2.xml")

    # === Build classpath
    jars = [os.path.join(lib_path, jar) for jar in os.listdir(lib_path) if jar.endswith(".jar")]
    jars.append(mpxj_jar)
    classpath = ";".join(jars)  # Windows uses ; separator

    # === Start JVM
    if not jpype.isJVMStarted():
        jpype.startJVM(jvm_path, "-ea", f"-Djava.class.path={classpath}",
        f"-Dlog4j.configurationFile={log_config_path}",
         convertStrings=True)

    # === Load Project
    MPPReader = jpype.JClass("org.mpxj.mpp.MPPReader")
    reader = MPPReader()
    project = reader.read(mpp_file)

    # === Task Data Extraction
    task_data = []
    for task in project.getTasks():
        if task.getName() is None:
            continue

        task_data.append({
        "ID": task.getID(),
        "Name": task.getName(),
        "Start": str(task.getStart()).split("T")[0] if task.getStart() else "",
        "Finish": str(task.getFinish()).split("T")[0] if task.getFinish() else "",
        "Duration": str(task.getDuration()),
        "% Complete": task.getPercentageComplete(),
        "Resource Names": ", ".join([
            r.getName()
            for a in task.getResourceAssignments()
            if (r := a.getResource()) and r.getName()
        ]) or "unassigned",

        "Working place": task.getText(3),   # Text3
        "Hours": task.getText(2),           # Text2
        "Unit": task.getText(4),            # Text4
        "Notes": task.getNotes()
        })

    # JSON processing
    result = {}
    ma_du_an = None
    for i in task_data:
        if i["ID"] == 0:
            ma_du_an = extract_project_code(i["Name"])
        if i["Resource Names"] != "unassigned":
            members = str(i["Resource Names"]).split(",")
            if (i["Working place"] != None ) and ( i["Hours"] != None):
                start = pd.to_datetime(i["Start"])
                end = pd.to_datetime(i["Finish"])
                days_task = (end - start).days + 1
                # weekend_days = amount_weekend_days(start, end)
                # working_days = days_task - weekend_days
                working_days = days_task
                QTY = round(float(i["Hours"])/working_days, 2) if working_days > 0 else 0
                task = {
                    "Mô tả công việc": i["Name"],
                    "Kế hoạch - Từ": i["Start"],
                    "Kế hoạch - Đến": i["Finish"],
                    "QTY": QTY,
                    "Nơi làm việc": i["Working place"]
                }

                for name in members:
                    name = name.strip()
                    if not name or name.lower() == "none":
                        continue
                    if name not in result:
                        result[name] = {}
                    if ma_du_an not in result[name]:
                        result[name][ma_du_an] = []
                    result[name][ma_du_an].append(task)

    # Processing output clean
    output = []
    for name, projects in result.items():
        if not name.strip():
            continue
        output.append({
            "Tên nhân sự": name,
            "Dự án": [
                {
                    "Mã dự án": code,
                    "Thông tin": tasks
                } for code, tasks in projects.items()
            ]
        })

    # Clean the output before returning and/or writing to JSON
    json_output = clean_json(output)

    # if json_output_path:
    #     with open(json_output_path, "w", encoding="utf-8") as f:
    #         json.dump(json_output, f, ensure_ascii=False, indent=2)
    print("json_output:", json_output)

    return json_output

def extract_project_code(name):
    if not name:
        return ""
    # Match code-like prefix (e.g. ES296-A2402)
    match = re.match(r"\s*([A-Z]+\d{3,}-[A-Z0-9]+)", name)
    return match.group(1) if match else name.strip()

def clean_json(data):
    if isinstance(data, dict):
        return {k: clean_json(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [clean_json(item) for item in data]
    else:
        return safe_convert(data)
    
def safe_convert(obj):
    if pd.isna(obj):  # Handles NaN and NaT
        return None
    if isinstance(obj, pd.Timestamp):
        return obj.strftime("%Y-%m-%d")
    return obj

def get_weekend_columns(df: pd.DataFrame) -> List[int]:
    """
    Hàm để xác định các cột là ngày cuối tuần trong DataFrame.
    Trả về danh sách các chỉ số cột tương ứng với ngày cuối tuần.
    """
    weekend_cols = []
    for col in df.columns:
        try:
            date = pd.to_datetime(col, errors='coerce')
            if date is not pd.NaT and date.weekday() in [5, 6]:
                weekend_cols.append(col)
        except:
            continue
    return weekend_cols

def save_file_minio(minio_client: Minio, file_path: str) -> str:
    try:
        bucket_name = "estec"
        object_name = f"data/POD/TimeTracker/Output/{file_path.split('/')[-1]}"
        minio_client.fput_object(bucket_name, object_name, file_path)
        output_path = f"{bucket_name}/{object_name}"
        # print(f"File saved to MinIO at {output_path}")
        return output_path
    except Exception as e:
        raise Exception(f"Error saving file to MinIO: {str(e)}")

# def merge_cells_by_columns(worksheet, df, col_indices):
#     for col in col_indices:
#         start_row = 2
#         for row in range(3, len(df) + 2):
#             if df.iloc[row - 2, col] != df.iloc[row - 3, col]:
#                 if row - 1 > start_row:
#                     worksheet.merge_cells(start_row=start_row, start_column=col + 1, end_row=row - 1, end_column=col + 1)
#                 start_row = row
#         # Merge cho phần cuối cùng
#         if len(df) + 1 > start_row:
#             worksheet.merge_cells(start_row=start_row, start_column=col + 1, end_row=len(df) + 1, end_column=col + 1)

def merge_cells_by_columns(worksheet, df, col_indices, name_col=1, codeproject_col=2):
    """
    worksheet: sheet excel openpyxl
    df: DataFrame
    col_indices: list các cột muốn merge (0-based index)
    name_col: cột "Tên nhân sự" (mặc định = 1)
    """
    for col in col_indices:
        start_row = 2
        for row in range(3, len(df) + 2):
            if col == codeproject_col:
                if df.iloc[row - 2, col] != df.iloc[row - 3, col]:
                    if row - 1 > start_row:
                        worksheet.merge_cells(start_row=start_row, start_column=col + 1, end_row=row - 1, end_column=col + 1)
                    start_row = row
            else:
                cur_val = df.iloc[row - 2, col]
                prev_val = df.iloc[row - 3, col]
                cur_name = df.iloc[row - 2, name_col]
                prev_name = df.iloc[row - 3, name_col]

                # nếu giá trị cột khác và tên cũng khác => ngắt nhóm
                if cur_val != prev_val and cur_name != prev_name:
                    if row - 1 > start_row:
                        worksheet.merge_cells(
                            start_row=start_row, start_column=col + 1,
                            end_row=row - 1, end_column=col + 1
                        )
                    start_row = row

        # merge đoạn cuối
        if len(df) + 1 > start_row:
            worksheet.merge_cells(
                start_row=start_row, start_column=col + 1,
                end_row=len(df) + 1, end_column=col + 1
            )

def check_overwork(df: pd.DataFrame) -> Optional[dict]:
    """
    Kiểm tra xem có công việc nào có QTY > 8 giờ hay không.
    Trả về dict chứa thông tin về nhân sự và ngày làm việc quá giờ.
    Nếu không có thì trả về None.
    """
    # Lấy các cột ngày
    date_cols = [col for col in df.columns if str(col).startswith('2025')]
    for col in date_cols:
        df[col] = df[col].astype(str).str.extract(r'(\d+)').astype(float).fillna(0)
    # Gom nhóm theo Tên nhân sự + cột ngày

    result = df.groupby('Tên nhân sự')[date_cols].sum().fillna(0)
    # print("Grouped result:", result)

    # Lọc chỗ > 8
    overwork = result[result > 8].dropna(how='all')

    # Schema
    output = []
    for name, row in overwork.iterrows():
        days = []
        for day, hours in row.items():
            if pd.notna(hours) and hours > 8:
                days.append({
                    "date_val": str(day),
                    "hours": float(hours)
                })
        if days:
            output.append({
                "employee": name,
                "overwork": days
            })

    if not output:
        return None  # Không có ai làm quá 8 giờ
    
    print("Overwork detected:", output)
    return output

def addfilter_columns(df: pd.DataFrame):
    cols = list(df.columns)
    # Vị trí 'Mô tả công việc'
    insert_at = cols.index("Mô tả công việc") + 1
    cols = (
        cols[:insert_at]
        + ["Tên nhân sự - filter", "Mã dự án - filter"]
        + cols[insert_at:]
    )
    df["Tên nhân sự - filter"] = df["Tên nhân sự"]
    df["Mã dự án - filter"] = df["Mã dự án"]
    df = df[cols]
    return df

def get_hours(cell_value) -> float:
    """
    Trích số giờ từ ô Excel, ví dụ:
    '8 | S' -> 8.0
    '12 | V' -> 12.0
    None, '' -> 0
    """
    if cell_value is None:
        return 0.0
    raw = str(cell_value).strip()
    if not raw:
        return 0.0
    try:
        return float(raw.split('|')[0].strip())
    except ValueError:
        return 0.0

def save_file_local(df: pd.DataFrame):
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"./minio/minio_data/POD/TimeTracker/Output/ES_{now}.xlsx"
    required_cols = ["Tên nhân sự - filter", "Mã dự án - filter"]
    if not all(col in df.columns for col in required_cols):
        df = addfilter_columns(df)
    else:
        df = df.drop(columns=required_cols)
        df = addfilter_columns(df) 
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=1, header=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Căn chỉnh các ô
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Tạo fill xanh nhạt cho header
        header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

        # Fill hồng cho cuối tuần
        weekend_fill = PatternFill(start_color="FADADD", end_color="FADADD", fill_type="solid")

        # Tạo border
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Tạo fill đỏ cho ô quá giờ
        overwork_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
        overwork_warning_fill = PatternFill(start_color="f7dc6f", end_color="f7dc6f", fill_type="solid")

        normal_fill = PatternFill(start_color="E2F0CB", end_color="E2F0CB", fill_type="solid")

        white_fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")

        # Xác định các cột là ngày cuối tuần
        weekend_cols = get_weekend_columns(df)
        overwork = check_overwork(df)
        
        print("Weekend columns:", weekend_cols)
        # Áp dụng cho các ô header
        for idx, col in enumerate(df.columns):
            cell = worksheet.cell(row=1, column=idx+1, value=col)
            if col in weekend_cols:
                cell.fill = weekend_fill
            else:
                cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment  
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=1+len(df), min_col=1, max_col=len(df.columns)), start=2):
            for cell in row:
                cell.alignment = center_alignment
                col_header = df.columns[cell.column - 1]
                if not str(col_header).startswith('2025'):
                    cell.border = border
                
                if str(col_header).startswith('2025'):
                    # Lấy tên nhân sự dòng này:
                    employee_name = df.iloc[row_idx - 2]['Tên nhân sự']
                    # Kiểm tra (employee_name, col_header) có trong overwork không?
                    overwork_flag = False
                    for item in overwork:
                        if item["employee"] == employee_name:
                            for day in item["overwork"]:
                                if str(day["date_val"]) == str(col_header):
                                    overwork_flag = True
                                    break
                        if overwork_flag:
                            break

                    val = get_hours(cell.value)
                    if val > 0:  # chỉ xử lý ô có giá trị
                        if val > 8:
                            cell.fill = overwork_fill
                        elif overwork_flag and val > 0:
                            cell.fill = overwork_warning_fill
                        elif val > 0:
                            cell.fill = normal_fill
                
        merge_cells_by_columns(worksheet, df, [0,1,2])
    print("✅ Xuất file thành công với ô đã được merge theo nhóm.")
    return filename, overwork

def generate_dataframe(json_data: List[dict], stt_max) -> str:
    # print("Generating output path...")
    tasks = json_data
    rows = []
    stt_global = stt_max
    # print("Processing tasks...")

    # Get range date
    start_dates = []
    end_dates = []
    for person in tasks:
        for project in person["Dự án"]:
            for task in project["Thông tin"]:
                if task.get("Kế hoạch - Từ"):
                    start_dates.append(task["Kế hoạch - Từ"])
                if task.get("Kế hoạch - Đến"):
                    end_dates.append(task["Kế hoạch - Đến"])

    date_min = min(datetime.strptime(d, "%Y-%m-%d") for d in start_dates)
    date_max = max(datetime.strptime(d, "%Y-%m-%d") for d in end_dates)
    print(f"Date range: {date_min} -> {date_max}")
    calendar_days = pd.date_range(start=date_min, end=date_max)

    for person in tasks:
        name = person["Tên nhân sự"]
        # print(f"Processing tasks for {name}...")
        stt_global += 1
        for project in person["Dự án"]:
            ma_du_an = project["Mã dự án"]
            for task in project["Thông tin"]:
                description = task["Mô tả công việc"]
                start = task["Kế hoạch - Từ"]
                end = task["Kế hoạch - Đến"]
                QTY = task["QTY"]
                place = task["Nơi làm việc"]
                
                row = {
                    "STT": stt_global,
                    "Tên nhân sự": name,
                    "Mã dự án": ma_du_an,
                    "Mô tả công việc": description,
                    "Thời gian bắt đầu": start,
                    "Thời gian kết thúc": end
                }

                start_date = datetime.strptime(start, "%Y-%m-%d")
                end_date = datetime.strptime(end, "%Y-%m-%d")

                for date in calendar_days:
                    if start_date <= date <= end_date:
                        if date.weekday() < 5:
                            row[date.strftime("%Y-%m-%d")] = f"{QTY} | {place}"
                        else:
                            row[date.strftime("%Y-%m-%d")] = f"{QTY} | {place}"
                    else:
                        row[date.strftime("%Y-%m-%d")] = None

                rows.append(row)
    # Tạo DataFrame
    df = pd.DataFrame(rows)
    return df

def amount_weekend_days(start_date: datetime, end_date: datetime) -> int:
    """
    Hàm để tính số ngày cuối tuần (Thứ 7, Chủ nhật) trong khoảng thời gian từ start_date đến end_date.
    Trả về số lượng ngày cuối tuần.
    """
    if not isinstance(start_date, datetime) or not isinstance(end_date, datetime):
        raise ValueError("start_date and end_date must be datetime objects")
    
    if start_date > end_date:
        raise ValueError("start_date must be before end_date")

    weekend_days = 0
    current_date = start_date

    while current_date <= end_date:
        if current_date.weekday() in [5, 6]:
            weekend_days += 1
        current_date += timedelta(days=1)
    print(f"Weekend days from {start_date} to {end_date}: {weekend_days}")
    return weekend_days

def processing_json(file_path: BytesIO): 
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    
    MaDuAn = sheet["M2"].value
    header = [cell.value for cell in sheet[7][:13]]
    sub_header = [cell.value for cell in sheet[8][:13]]

    # print(f"MaDuAn: {MaDuAn}")
    # print(f"header: {header}")
    # print(f"sub_header: {sub_header}")

    Header = []
    previous_h1 = ""
    for h1, h2 in zip(header, sub_header):
        if h1 is None:
            h1 = previous_h1
        else:
            previous_h1 = h1
        h2 = "" if h2 is None or "#REF" in str(h2) else h2
        Header.append(f"{h1} - {h2}" if h2 else h1)
    
    # print(f"Header: {Header}")

    data = sheet.iter_rows(min_row=9, min_col=1, max_col=13, values_only=True)
    df = pd.DataFrame(data, columns=Header).dropna(subset=[Header[0]])
    df = df[:-1]

    result = {}
    message = []
    for i in df.index:
        row = df.loc[i]
        if pd.isna(row[Header[3]]):
            continue
        members = str(row[Header[3]]).split(",")

        start = pd.to_datetime(row[Header[5]])
        end = pd.to_datetime(row[Header[6]])
        days_task = (end - start).days + 1

        # days_task = (row[Header[6]] - row[Header[5]]).days + 1
        # weekend_days = amount_weekend_days(row[Header[5]], row[Header[6]])
        # working_days = days_task - weekend_days
        working_days = days_task

        # print(f"Days task: {days_task}")

        # if row[Header[1]]:
        #     MoTaCongViec = row[Header[1]]  
        # else:
        #     MoTaCongViec = "Không có mô tả công việc"
        #     message.append(f'''Dự án {MaDuAn} có công việc không có mô tả công việc. Vui lòng kiểm tra lại.''')

        # if isinstance(row[Header[5]], datetime):
        #     KeHoachTu = row[Header[5]].strftime("%Y-%m-%d")
        # else:
        #     KeHoachTu = None
        #     message.append(f'''Dự án {MaDuAn} có công việc {MoTaCongViec}: Không có ngày bắt đầu. Vui lòng kiểm tra lại.''')
            
        # if isinstance(row[Header[6]], datetime):
        #     KeHoachDen = row[Header[6]].strftime("%Y-%m-%d")
        # else:
        #     KeHoachDen = None
        #     message.append(f'''Dự án {MaDuAn} có công việc {MoTaCongViec}: Không có ngày kết thúc. Vui lòng kiểm tra lại.''')

        MoTaCongViec = row[Header[1]] if row[Header[1]] else "Không có mô tả công việc"
        KeHoachTu = row[Header[5]].strftime("%Y-%m-%d") if isinstance(row[Header[5]], datetime) else "Không có ngày bắt đầu"
        KeHoachDen = row[Header[6]].strftime("%Y-%m-%d") if isinstance(row[Header[6]], datetime) else "Không có ngày kết thúc"
        print(f"MaDuAn: {MaDuAn}, MoTaCongViec: {MoTaCongViec}, KeHoachTu: {KeHoachTu}, KeHoachDen: {KeHoachDen}")
        # print(f"working_days: {working_days}, days_task: {days_task}, weekend_days: {weekend_days}")
        # error_message = []
        # message = ""
        
        # if working_days <= 0:
            # return {
            #     "status": "error",
            #     "message": f'''Dự án {MaDuAn} có công việc {MoTaCongViec}: Có ngày làm việc rơi vào Thứ 7 hoặc Chủ nhật. Vui lòng điều chỉnh lại ngày bắt đầu & kết thúc. Ngày bắt đầu hiện tại: {KeHoachTu}, Ngày kết thúc hiện tại: {KeHoachDen}.'''
            # }
            # message += f'''Dự án {MaDuAn} có công việc {MoTaCongViec}: Có ngày làm việc rơi vào Thứ 7 hoặc Chủ nhật. Vui lòng điều chỉnh lại ngày bắt đầu & kết thúc. Ngày bắt đầu hiện tại: {KeHoachTu}, Ngày kết thúc hiện tại: {KeHoachDen}.\n'''
            # message.append(f'''Dự án {MaDuAn} có công việc {MoTaCongViec}: Có ngày làm việc rơi vào Thứ 7 hoặc Chủ nhật. Vui lòng điều chỉnh lại ngày bắt đầu & kết thúc. Ngày bắt đầu hiện tại: {KeHoachTu}, Ngày kết thúc hiện tại: {KeHoachDen}.''')
        QTY = round(float(row[Header[8]])/working_days, 2) if working_days > 0 else 0
        NoiLamViec = row[Header[7]] if row[Header[7]] else "Không có nơi làm việc"
        # if QTY > 8:
        #     return {
        #         "status": "error",
        #         "message": f'''Dự án {MaDuAn} có công việc {MoTaCongViec}: QTY ({QTY}) giờ > 8 giờ. Vui lòng điều chỉnh lại ngày bắt đầu và kết thúc. Ngày bắt đầu hiện tại: {KeHoachTu}, Ngày kết thúc hiện tại: {KeHoachDen}.'''
        #     }
        # if QTY <= 0:
        #     return {
        #         "status": "error",
        #         "message": f'''Dự án {MaDuAn} có công việc {MoTaCongViec}: QTY ({QTY}) giờ <= 0 giờ. Vui lòng điều chỉnh lại ngày bắt đầu và kết thúc. Ngày bắt đầu hiện tại: {KeHoachTu}, Ngày kết thúc hiện tại: {KeHoachDen}.'''
        #     }
        #     message += f'''Dự án {MaDuAn} có công việc {MoTaCongViec}: QTY ({QTY}) giờ <= 0 giờ. Vui lòng điều chỉnh lại ngày bắt đầu và kết thúc. Ngày bắt đầu hiện tại: {KeHoachTu}, Ngày kết thúc hiện tại: {KeHoachDen}.\n'''
        #     message.append(f'''Dự án {MaDuAn} có công việc {MoTaCongViec}: QTY ({QTY}) giờ <= 0 giờ. Vui lòng điều chỉnh lại ngày bắt đầu và kết thúc. Ngày bắt đầu hiện tại: {KeHoachTu}, Ngày kết thúc hiện tại: {KeHoachDen}.''')
        
        task = {
            "Mô tả công việc": MoTaCongViec,
            "Kế hoạch - Từ": KeHoachTu,
            "Kế hoạch - Đến": KeHoachDen,
            "QTY": QTY,
            "Nơi làm việc": NoiLamViec
        }

        for name in members:
            name = name.strip()
            if not name or name.lower() == "none":
                continue
            if name not in result:
                result[name] = {}
            if MaDuAn not in result[name]:
                result[name][MaDuAn] = []
            result[name][MaDuAn].append(task)

    if message:
        return {
            "status": "error",
            "message": message
        }

    output_json = []
    for name, projects in result.items():
        if not name.strip():
            continue
        output_json.append({
            "Tên nhân sự": name,
            "Dự án": [
                {
                    "Mã dự án": code,
                    "Thông tin": tasks
                } for code, tasks in projects.items()
            ]
        })
    # print(output_json)
    return output_json
