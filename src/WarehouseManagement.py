import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from psycopg2 import sql
from minio import Minio
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from psycopg2.extras import RealDictCursor
import io
import openpyxl

# def WarehouseStatistical_View_function(input, conn):
#     try:
#         # print(input)
#         cursor = conn.cursor()
#         query = """ 
#             SELECT * FROM "WS_Statistical"
#             WHERE 1=1
#         """
#         params = []

#         if input.filter.part_no:
#             query += " AND \"part_no\" ILIKE %s"
#             params.append(f"%{input.filter.part_no}%")

#         if input.filter.origin:
#             query += " AND \"origin\" ILIKE %s"
#             params.append(f"%{input.filter.origin}%")

#         if input.filter.seri_number:
#             query += " AND \"seri_number\" ILIKE %s"
#             params.append(f"%{input.filter.seri_number}%")

#         query += " ORDER BY \"time\" DESC LIMIT 1000"

#         cursor.execute(query, params)
#         rows = cursor.fetchall()
#         items = []
#         for row in rows:
#             # --- Lấy tổng nhập và tổng xuất cho part_no hiện tại ---
#             cursor2 = conn.cursor()
#             cursor2.execute("""
#                 SELECT 
#                     COALESCE(SUM("quantity"), 0)
#                 FROM "WS_Import"
#                 WHERE "part_no" = %s AND "deleted" = FALSE
#             """, (row[4],))
#             total_import = cursor2.fetchone()[0]

#             cursor2.execute("""
#                 SELECT 
#                     COALESCE(SUM("quantity"), 0)
#                 FROM "WS_Export"
#                 WHERE "part_no" = %s AND "deleted" = FALSE
#             """, (row[4],))
#             total_export = cursor2.fetchone()[0]
#             cursor2.close()
#             # --- Tính tồn kho ---
#             stock_quantity = total_import - total_export

#             item = {
#                 "id": row[0],
#                 "product_name": row[1],
#                 "description": row[2],
#                 "time": row[3],
#                 "part_no": row[4],
#                 "origin": row[5],
#                 "unit": row[6],
#                 "quantity_import": total_import,
#                 "quantity_export": total_export,
#                 "quantity_stock": row[7] + stock_quantity,
#                 "seri_number": row[8],
#                 "location": row[9],
#                 "entered_by": row[10],
#                 "status": row[11],
#             }
#             items.append(item)
        
#         # print(len(items))

#         start_idx = (input.pagination - 1) * input.page_size
#         end_idx = input.pagination * input.page_size
#         items = items[start_idx:end_idx]
#         # print(len(items))
#         return {
#             "status": "success",
#             "data": items
#         }
#     except Exception as e:
#         return {
#             "status": "error",
#             "message": str(e)
#         }
#     finally:
#         cursor.close()

def WarehouseStatistical_View_function(input, conn):
    try:
        # print(input)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        query = """ 
            SELECT 
                i.*, 
                s."product_name", 
                s."description", 
                s."origin", 
                s."entered_by", 
                s."unit", 
                s."location",
                COALESCE(e."quantity", 0) AS quantity_export
            FROM "WS_Import" i
            LEFT JOIN "WS_Statistical" s
                ON i."statistical_id" = s."id"
            LEFT JOIN "WS_Export" e 
                ON i."seri_number" = e."seri_number" AND e."deleted" = FALSE
            WHERE i."deleted" = FALSE
        """
        params = []

        if input.filter.part_no:
            query += " AND i.\"part_no\" ILIKE %s"
            params.append(f"%{input.filter.part_no}%")

        if input.filter.origin:
            query += " AND s.\"origin\" ILIKE %s"
            params.append(f"%{input.filter.origin}%")

        if input.filter.seri_number:
            query += " AND i.\"seri_number\" ILIKE %s"
            params.append(f"%{input.filter.seri_number}%")

        if input.filter.project_code:
            query += " AND i.\"project_code\" ILIKE %s"
            params.append(f"%{input.filter.project_code}%")

        if input.filter.datetime_import:
            # Giả sử input.filter.datetime_import là chuỗi "2025-10-17"
            import_date = input.filter.datetime_import.date()
            query += " AND i.\"time\"::date = %s::date"
            params.append(import_date)

        else:
            query += " ORDER BY i.\"time\" DESC LIMIT 1000"

        cursor.execute(query, params)
        rows = cursor.fetchall()
        items = []
        for row in rows:
            item = {
                "id": row.get("id", None),
                "product_name": row.get("product_name", ""),
                "description": row.get("description", ""),
                "time": row.get("time", ""),
                "project_code": row.get("project_code", ""),
                "part_no": row.get("part_no", ""),
                "origin": row.get("origin", ""),
                "unit": row.get("unit", ""),
                "quantity_import": row.get("quantity", ""),
                "quantity_export": row.get("quantity_export", ""),
                "quantity_stock": row.get("quantity", "") - row.get("quantity_export", ""),
                "seri_number": row.get("seri_number", ""),
                "location": row.get("location", ""),
                "entered_by": row.get("entered_by", ""),
                "status": row.get("status", 0),
                "manufacturing_date": row.get("manufacturing_date", None)
            }
            items.append(item)
        
        # print(len(items))

        start_idx = (input.pagination - 1) * input.page_size
        end_idx = input.pagination * input.page_size
        items = items[start_idx:end_idx]
        # print(len(items))
        return {
            "status": "success",
            "data": items
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }
    finally:
        cursor.close()

def WarehouseStatistical_View_Detail_function(input, conn):
    try:
        cursor = conn.cursor()
        query = """ 
            SELECT * FROM "WS_Statistical" WHERE id = %s
        """
        cursor.execute(query, (input.id,))
        row = cursor.fetchone()
        if row:
            item = {
                "id": row[0],
                "product_name": row[1],
                "description": row[2],
                "time": row[3],
                "part_no": row[4],
                "origin": row[5],
                "unit": row[6],
                "quantity": row[7],
                "seri_number": row[8],
                "location": row[9],
                "entered_by": row[10],
                "status": row[11]
            }
            return {
                "status": "success",
                "data": item
            }
        else:
            return {
                "status": "error",
                "message": f"Item with ID {input.id} not found."
            }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }
    finally:
        cursor.close()

def WarehouseStatistical_DML_Insert_function(input, conn):
    try:
        cursor = conn.cursor()
        query = """ 
            INSERT INTO "WS_Statistical" 
            ("product_name", "description", "time", "part_no", "origin", "unit", "quantity", "seri_number", "location", "entered_by", "status") 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) 
            RETURNING id
        """
        cursor.execute(query, (
            input.form.product_name,
            input.form.description,
            input.form.time,
            input.form.part_no,
            input.form.origin,
            input.form.unit,
            input.form.quantity,
            input.form.seri_number,
            input.form.location,
            input.form.entered_by,
            input.form.status
        ))
        new_id = cursor.fetchone()[0]
        conn.commit()
        return {
            "status": "success",
            "message": f"Đã thêm danh mục ID {new_id}.",
            "ID": new_id
        }
    except Exception as e:
        conn.rollback()
        return {
            "status": "error",
            "message": str(e)
        }
    finally:
        cursor.close()

def WarehouseStatistical_DML_Update_function(input, conn):
    try:
        cursor = conn.cursor()
        query = """ 
            UPDATE "WS_Statistical" 
            SET "product_name" = %s, "description" = %s, "time" = %s, "part_no" = %s, "origin" = %s, "unit" = %s, "quantity" = %s, "seri_number" = %s, "location" = %s, "entered_by" = %s, "status" = %s 
            WHERE id = %s
        """
        cursor.execute(query, (
            input.form.product_name,
            input.form.description,
            input.form.time,
            input.form.part_no,
            input.form.origin,
            input.form.unit,
            input.form.quantity,
            input.form.seri_number,
            input.form.location,
            input.form.entered_by,
            input.form.status,
            input.form.id
        ))
        if cursor.rowcount == 0:
            return {
                "status": "error",
                "message": f"ID {input.form.id} không tồn tại."
            }
        conn.commit()
        return {
            "status": "success",
            "message": f"Đã cập nhật danh mục ID {input.form.id}."
        }
    except Exception as e:
        conn.rollback()
        return {
            "status": "error",
            "message": str(e)
        }
    finally:
        cursor.close()

def WarehouseStatistical_DML_Delete_function(input, conn):
    try:
        cursor = conn.cursor()
        query = """ 
            DELETE FROM "WS_Statistical" WHERE id = %s
        """
        cursor.execute(query, (input.form.id,))
        if cursor.rowcount == 0:
            return {
                "status": "error",
                "message": f"ID {input.form.id} không tồn tại."
            }
        conn.commit()
        return {
            "status": "success",
            "message": f"Đã xóa danh mục ID {input.form.id}."
        }
    except Exception as e:
        conn.rollback()
        return {
            "status": "error",
            "message": str(e)
        }
    finally:
        cursor.close()

def WarehouseStatistical_Upload_function(conn, file):
    try:
        filename = file.filename
        ext = os.path.splitext(filename)[-1].lower()

        # Đọc file Excel/CSV
        if ext == ".xlsx":
            df = pd.read_excel(file.file, engine="openpyxl", header=0) 
        elif ext == ".xls":
            df = pd.read_excel(file.file, engine="xlrd", header=0)
        elif ext == ".csv":
            try:
                df = pd.read_csv(file.file, encoding="utf-8", header=0)
            except UnicodeDecodeError:
                df = pd.read_csv(file.file, encoding="latin1", header=0)
        else:
            return {"status": "error", "message": "Định dạng file không được hỗ trợ"}

        # Chuẩn hóa dữ liệu
        df = df.replace({np.nan: None})

        # Map lại cột Excel -> DB (tuỳ thuộc file Excel của bạn)
        df = df.rename(columns={
            "Tên sản phẩm": "product_name",
            "Thông tin sản phẩm": "description",
            "Ngày tạo": "time",
            "Mã sản phẩm": "part_no",
            "Nhà sản xuất": "origin",
            "Đơn vị": "unit",
            "Seri sản phẩm": "seri_number",
            "Người tạo": "entered_by"
        })

        # Thêm cột auto
        # if "time" in df.columns:
        #     df["time"] = pd.to_datetime(df["time"], dayfirst=True, errors="coerce")
        #     df["time"] = df["time"].where(df["time"].notnull(), None)
        # else:
        #     df["time"] = datetime.now()

        df["quantity"] = 1
        df["location"] = None
        df["status"] = 1
        df["manufacturing_date"] = None
        df["time"] = pd.to_datetime(df["time"], format="%d/%m/%Y", errors="coerce")
        df["time"] = df["time"].ffill()

        print(df["time"])
        print(df["time"].head().apply(lambda x: type(x)))
        print(df["time"].tail().apply(lambda x: type(x)))
    
        # Insert vào DB
        with conn.cursor() as cursor:
            for _, row in df.iterrows():
                cursor.execute("""
                    INSERT INTO "WS_Statistical"
                    (product_name, description, time, part_no, origin, unit, quantity, seri_number, location, entered_by, status, manufacturing_date)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    row["product_name"],
                    row["description"],
                    row["time"],
                    row["part_no"],
                    row["origin"],
                    row["unit"],
                    row["quantity"],
                    row["seri_number"],
                    row["location"],
                    row["entered_by"],
                    row["status"],
                    row["manufacturing_date"]
                ))
            conn.commit()

        return {
            "status": "success",
            "message": "Import dữ liệu thành công"
        }

    except Exception as e:
        return {
            "status": "error", 
            "message": str(e)
        }
    # finally:
    #     cursor.close()

### Warehouse - Import, Export ###

def WarehouseImportExport_View_function(conn, table_name: str):
    try:
        cursor = conn.cursor()

        if table_name not in ["WS_Import", "WS_Export"]:
            return {
                "status": "error",
                "message": f"Bảng '{table_name}' không được phép truy cập."
            }

        query = f'''
            SELECT * 
            FROM "{table_name}" 
            WHERE "deleted" = FALSE
            ORDER BY "import_time" DESC
            '''
        cursor.execute(query)
        rows = cursor.fetchall()

        items = []
        for row in rows:
            item = {
                "id": row[0],
                f"{'import' if table_name == 'WS_Import' else 'export'}_id": row[1],
                "time": row[2],
                f"{'import' if table_name == 'WS_Import' else 'export'}_time": row[3],
                "project_code": row[4],
                "product_name": row[5],
                "part_no": row[6],
                "origin": row[7],
                "quantity": row[8],
                "seri_number": row[9]
            }
            items.append(item)

        return {
            "status": "success",
            "data": items
        }

    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }
    finally:
        cursor.close()

def WarehouseImportExport_View_Detail_function(input, conn, table_name: str):
    try:
        cursor = conn.cursor()

        query = sql.SQL("""
            SELECT * 
            FROM {table} 
            WHERE id = %s AND "deleted" = FALSE
        """).format(table=sql.Identifier(table_name))
        cursor.execute(query, (input.id,))
        row = cursor.fetchone()

        if row:
            item = {
                "id": row[0],
                f"{'import' if table_name == 'WS_Import' else 'export'}_id": row[1],
                "time": row[2],
                f"{'import' if table_name == 'WS_Import' else 'export'}_time": row[3],
                "project_code": row[4],
                "product_name": row[5],
                "part_no": row[6],
                "origin": row[7],
                "quantity": row[8],
                "seri_number": row[9]
            }
            return {"status": "success", "data": item}
        else:
            return {
                "status": "error",
                "message": f"Item with ID {input.id} not found."
            }

    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        cursor.close()

def WarehouseImportExport_DML_Insert_function(input, conn, option):
    try:
        cursor = conn.cursor()
        
        if option == "import":
            table_name = "WS_Import"
            id_field = "import_id"
            id_value = input.form.ticket_id
            time_field = "import_time"
            time_value = input.form.ticket_time
        elif option == "export":
            table_name = "WS_Export"
            id_field = "export_id"
            id_value = input.form.ticket_id
            time_field = "export_time"
            time_value = input.form.ticket_time
        else:
            return {
                "status": "error", 
                "message": "Option không hợp lệ. Chỉ hỗ trợ 'import' hoặc 'export'."
                }
        
        query = sql.SQL("""
            INSERT INTO {table} 
            ({id_field}, "time", {time_field}, "project_code", "product_name", "part_no", "origin", "quantity", "seri_number")
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """).format(
            table=sql.Identifier(table_name),
            id_field=sql.Identifier(id_field),
            time_field=sql.Identifier(time_field)
        )
        
        cursor.execute(query, (
            id_value,
            input.form.time,
            time_value,
            input.form.project_code,
            input.form.product_name,
            input.form.part_no,
            input.form.origin,
            input.form.quantity,
            input.form.seri_number,
        ))

        new_id = cursor.fetchone()[0]
        conn.commit()

        return {
            "status": "success",
            "message": f"Đã thêm danh mục ID {new_id}.",
            "ID": new_id
        }

    except Exception as e:
        conn.rollback()
        return {
            "status": "error",
            "message": str(e)
        }

    finally:
        cursor.close()

def WarehouseImportExport_DML_Update_function(input, conn, option):
    # try:
        cursor = conn.cursor()
        print(input.form.ticket_time)

        if option == "import":
            table_name = "WS_Import"
            id_field = "import_id"
            id_value = str(input.form.ticket_id)
            time_field = "import_time"
            time_value = input.form.ticket_time if input.form.ticket_time else None
        elif option == "export":
            table_name = "WS_Export"
            id_field = "export_id"
            id_value = str(input.form.ticket_id)
            time_field = "export_time"
            time_value = input.form.ticket_time if input.form.ticket_time else None
        else:
            return {
                "status": "error", 
                "message": "Option không hợp lệ. Chỉ hỗ trợ 'import' hoặc 'export'."
                }
        print("ticket_id:", id_value)
        query = sql.SQL("""
            UPDATE {table}
            SET 
                {id_field} = %s, 
                "time" = %s, 
                {time_field} = %s, 
                "project_code" = %s, 
                "product_name" = %s, 
                "part_no" = %s, 
                "origin" = %s, 
                "quantity" = %s, 
                "seri_number" = %s
            WHERE id = %s
        """).format(
            table=sql.Identifier(table_name),
            id_field=sql.Identifier(id_field),
            time_field=sql.Identifier(time_field)
        )

        cursor.execute(query, (
            str(id_value),
            input.form.time,
            time_value,
            input.form.project_code,
            input.form.product_name,
            input.form.part_no,
            input.form.origin,
            input.form.quantity,
            input.form.seri_number,
            input.form.id
        ))

        if cursor.rowcount == 0:
            return {
                "status": "error",
                "message": f"ID {input.form.id} không tồn tại."
            }

        conn.commit()
        return {
            "status": "success",
            "message": f"Đã cập nhật danh mục ID {input.form.id}."
        }

    # except Exception as e:
    #     conn.rollback()
    #     return {
    #         "status": "error",
    #         "message": str(e)
    #     }

    # finally:
    #     cursor.close()

def WarehouseImportExport_DML_Delete_function(input, conn, option):
    try:
        cursor = conn.cursor()

        if option == "import":
            table_name = "WS_Import"
        elif option == "export":
            table_name = "WS_Export"
        else:
            return {
                "status": "error", 
                "message": "Option không hợp lệ. Chỉ hỗ trợ 'import' hoặc 'export'."
                }

        query = sql.SQL("""
            UPDATE {table} 
            SET
                "deleted" = TRUE
            WHERE "id" = %s
        """).format(
            table=sql.Identifier(table_name)
        )

        cursor.execute(query, (input.form.id,))

        if cursor.rowcount == 0:
            return {
                "status": "error",
                "message": f"ID {input.form.id} không tồn tại."
            }
        conn.commit()
        return {
            "status": "success",
            "message": f"Đã xóa danh mục ID {input.form.id}."
        }
    except Exception as e:
        conn.rollback()
        return {
            "status": "error",
            "message": str(e)
        }
    finally:
        cursor.close()

def WarehouseImportExport_Upload_function(conn, file, option: str):
    cursor = None
    try:
        filename = file.filename
        ext = os.path.splitext(filename)[1].lower()

        if ext == ".xlsx":
            df = pd.read_excel(file.file, engine="openpyxl", header=1)
        elif ext == ".xls":
            df = pd.read_excel(file.file, engine="xlrd", header=1)
        elif ext == ".csv":
            df = pd.read_csv(file.file, encoding="latin1", header=1)
        else:
            return {"status": "error", "message": "Định dạng file không được hỗ trợ."}
        print(df)
        df = df.replace({np.nan: None})
        df = df.where(pd.notnull(df), None)

        df = df.rename(columns={
            "Thời gian": "time",
            "Mã Dự án": "project_code",
            "Tên hàng": "product_name",
            "Mã hàng": "part_no",
            "Hãng": "origin",
            "Số lượng": "quantity",
            "Seri No.": "seri_number"
        })

        if "time" in df.columns:
            df["time"] = pd.to_datetime(df["time"], errors="coerce")
            df["time"] = df["time"].replace({pd.NaT: None})

        now = datetime.now()
        if option == "import":
            table_name = "WS_Import"
            if "import_time" in df.columns:
                df["import_time"] = pd.to_datetime(df["import_time"], errors="coerce")
                df["import_time"] = df["import_time"].replace({pd.NaT: None})
            else:
                df["import_time"] = now  
            id_col = "import_id"
            time_col = "import_time"
        elif option == "export":
            table_name = "WS_Export"
            if "export_time" in df.columns:
                df["export_time"] = pd.to_datetime(df["export_time"], errors="coerce")
                df["export_time"] = df["export_time"].replace({pd.NaT: None})
            else:
                df["export_time"] = now  
            id_col = "export_id"
            time_col = "export_time"
        else:
            return {"status": "error", "message": f"Tùy chọn '{option}' không hợp lệ."}
        
        # Nhân bản theo số lượng bảng ghi
        if option == "export":
            expanded_rows = []
            for _, row in df.iterrows():
                qty = int(row["quantity"]) if row["quantity"] else 1
                for _ in range(qty):
                    new_row = row.copy()
                    new_row["quantity"] = 1
                    expanded_rows.append(new_row)
            df = pd.DataFrame(expanded_rows)
            df = df.set_index("STT")
            df = df.reset_index(drop=True)
        print("df:", df.columns)


        datetime_columns = ["time", time_col]
        for col in datetime_columns:
            if col in df.columns:
                df[col] = df[col].replace({pd.NaT: None})
        
        print("df:", df)

        cursor = conn.cursor()
        insert_query = f"""
            INSERT INTO "{table_name}"
            ({id_col}, "time", {time_col}, "project_code", "product_name", "part_no", "origin", "quantity", "seri_number")
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """
        for _, row in df.iterrows():       
            values = (
                row.get(id_col),  
                row.get("time"),         
                datetime.now(),   
                row.get("project_code"),
                row.get("product_name"), 
                row.get("part_no"),
                row.get("origin"),
                row.get("quantity"),
                row.get("seri_number"),
            )
            print("values:", values)
            cursor.execute(insert_query, values)

        conn.commit()
        return {"status": "success", "message": f"Tải dữ liệu {option} thành công."}

    except Exception as e:
        conn.rollback()
        return {
            "status": "error", 
            "message": str(e)}
    finally:
        if cursor:
            cursor.close()

def WarehouseImportExport_Download_function(conn, input, minio_client: Minio, MINIO_BUCKET: str):
    # SốTT	Thời gian	Mã Dự án	Tên hàng	Mã hàng	Hãng	Số lượng	Seri No.

    cursor = None
    try:
        cursor = conn.cursor()
        # --- Chọn bảng phù hợp ---
        if input.option == "import":
            if input.ticket_id:
                query = '''
                    SELECT * FROM "WS_Import"
                    WHERE "import_id" = %s 
                '''
                cursor.execute(query, (input.ticket_id,))
            elif input.project_code:
                query = '''
                    SELECT * FROM "WS_Import"
                    WHERE "project_code" = %s 
                '''
                cursor.execute(query, (input.project_code,))

            elif input.ticket_id and input.project_code:
                query = '''
                    SELECT * FROM "WS_Import"
                    WHERE "import_id" = %s 
                    AND "project_code" = %s 
                '''
                cursor.execute(query, (input.ticket_id, input.project_code))
            object_prefix = "Import"
        elif input.option == "export":
            if input.ticket_id:
                query = '''
                    SELECT * FROM "WS_Export"
                    WHERE "export_id" = %s 
                '''
                cursor.execute(query, (input.ticket_id,))
            elif input.project_code:
                query = '''
                    SELECT * FROM "WS_Export"
                    WHERE "project_code" = %s 
                '''
                cursor.execute(query, (input.project_code))
            elif input.ticket_id and input.project_code:
                query = '''
                    SELECT * FROM "WS_Export"
                    WHERE "export_id" = %s 
                    AND "project_code" = %s 
                '''
                cursor.execute(query, (input.ticket_id, input.project_code))
            object_prefix = "Export"
        else:
            raise ValueError("Invalid option. Must be 'import' or 'export'.")

        # --- Lấy dữ liệu ---
        data = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        df = pd.DataFrame(data, columns=columns)
        print("df:",df)
        columns_file = ["STT", "Thời gian", "Mã Dự án", "Tên hàng", "Mã hàng", "Hãng", "Số lượng", "Seri No."]
        df["STT"] = range(1, len(df) + 1)
        df["Thời gian"] = df["time"]
        df["Mã Dự án"] = df["project_code"]
        df["Tên hàng"] = df["product_name"]
        df["Mã hàng"] = df["part_no"]
        df["Hãng"] = df["origin"]
        df["Số lượng"] = df["quantity"]
        df["Seri No."] = df["seri_number"]
        df = df[columns_file]
        # --- Ghi ra file tạm ---
        if input.ticket_id:
            filename = f"{object_prefix}_{input.ticket_id}.xlsx"
        elif input.project_code:
            filename = f"{object_prefix}_{input.project_code}.xlsx"
        path_file = f"./minio/minio_data/Workshop/Warehouse/{object_prefix}/{filename}"
        os.makedirs(os.path.dirname(path_file), exist_ok=True)
        with pd.ExcelWriter(path_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1", startrow=1)
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]

            # --- Phieu nhap / phieu xuat ----
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            cell_title = worksheet.cell(row=1, column=1)
            if object_prefix == "Import":
                cell_title.value = f"PHIẾU NHẬP"
            else:
                cell_title.value = f"PHIẾU XUẤT"
            cell_title.font = Font(size=18, bold=True)
            cell_title.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell_title.alignment = Alignment(horizontal="center", vertical="center")

            # --- Style header dữ liệu (dòng 2) ---
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))

            for col_idx, col in enumerate(df.columns, 1):
                cell = worksheet.cell(row=2, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # --- Border cho tất cả ô dữ liệu ---
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row,
                                        min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border

        # --- Upload lên MinIO ---
        object_name = f"data/Workshop/Warehouse/{object_prefix}/{filename}"
        with open(path_file, "rb") as file_data:
            file_stat = os.stat(path_file)
            minio_client.put_object(
                MINIO_BUCKET,
                object_name,
                file_data,
                length=file_stat.st_size,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # --- Xóa file local sau khi upload ---
        # os.remove(path_file)

        # --- Tạo URL tải file ---
        url = minio_client.presigned_get_object(
            MINIO_BUCKET,
            object_name,
            expires=timedelta(hours=1)
        )

        return {
            "status": "success",
            "url": url
        }

    except Exception as e:
        conn.rollback()
        return {
            "status": "error",
            "message": str(e)
        }

    finally:
        if cursor:
            cursor.close()

# ------------------------
# Warehouse_Installation
# ------------------------
def WarehouseInstallation_Upload_function(conn, owner, file):
    try:
        filename = file.filename.rsplit('.', 1)[0]
        project_code = filename.split('-')[:2]
        project_code = '-'.join(project_code)

        cursor = conn.cursor()
        contents = file.file.read()
        df = pd.read_excel(io.BytesIO(contents))
        for _, row in df.iterrows():
            serial_number = row.get("SERIAL NUMBER")
            # Nếu serial_number là NaN hoặc chuỗi rỗng, chuyển thành None
            if pd.isna(serial_number) or serial_number == "":
                serial_number = None

            query = """
                INSERT INTO "WS_Installation" (
                    "id",
                    "higher_lever_function",
                    "location",      
                    "dt",
                    "quantity",            
                    "description",                      
                    "part_no",               
                    "seri_number",
                    "manufacturer",
                    "project_code"
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(query, (
                row.get("NO."),
                row.get("HIGHER LEVEL FUNCTION"),
                row.get("LOCATION"),
                row.get("DT"),
                row.get("QUANTITY"),
                row.get("DESCRIPTION 1"),
                row.get("ORDER NUMBER"),
                serial_number,
                row.get("MANUFACTURER"),
                project_code
            ))

        conn.commit()
        cursor.close()

        return {
            "status": "success",
            "message": f"Tải lên thành công {len(df)} dòng dữ liệu."
        }

    except Exception as e:
        conn.rollback()
        return {
            "status": "error",
            "message": str(e)
        }

def WarehouseInstallation_Download_function(conn, input, minio_client: Minio, MINIO_BUCKET: str):
    cursor = None
    try:
        cursor = conn.cursor()
        query = '''
            SELECT * FROM "WS_Installation"
            WHERE 1=1
        '''
        params = []
        if input.project_code:
            query += ' AND "project_code" = %s '
            params.append(input.project_code)
        if input.cabinet_no:
            query += ' AND "cabinet_no" = %s '
            params.append(input.cabinet_no)

        cursor.execute(query, params)
        data = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        df = pd.DataFrame(data, columns=columns)

        columns_file = [
            "NO.", "HIGHER LEVEL FUNCTION", "LOCATION", "DT", "QUANTITY",
            "DESCRIPTION 1", "ORDER NUMBER", "SERIAL NUMBER", "MANUFACTURER"
        ]
        df["NO."] = df["id"]
        df["HIGHER LEVEL FUNCTION"] = df["higher_lever_function"].apply(
            lambda x: f"'{x}" if isinstance(x, str) and x.startswith('=') else x
        )
        df["LOCATION"] = df["location"]
        df["DT"] = df["dt"]
        df["QUANTITY"] = df["quantity"]
        df["DESCRIPTION 1"] = df["description"]
        df["ORDER NUMBER"] = df["part_no"]
        df["SERIAL NUMBER"] = df["seri_number"]
        df["MANUFACTURER"] = df["manufacturer"]
        if not input.project_code:
            input.project_code = df["project_code"].iloc[0]
        if not input.cabinet_no:
            input.cabinet_no = df["cabinet_no"].iloc[0]

        # --- Tạo file Excel ---
        path_file = f"./minio/minio_data/Workshop/Warehouse/Installation/{input.project_code}-PARTLIST-{input.cabinet_no}.xlsx"
        df = df[columns_file]
        df.to_excel(path_file, index=False)

        # --- Tạo workbook ---
        workbook = openpyxl.load_workbook(path_file)
        worksheet = workbook.active
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, str) and "\n" in cell.value:
                    cell.alignment = Alignment(wrapText=True, vertical="top")
        workbook.save(path_file)
        # --- Upload lên MinIO ---
        object_name = f"data/Workshop/Warehouse/Installation/{input.project_code}-PARTLIST-{input.cabinet_no}.xlsx"
        with open(path_file, "rb") as file_data:
            file_stat = os.stat(path_file)
            minio_client.put_object(
                MINIO_BUCKET,
                object_name,
                file_data,
                length=file_stat.st_size,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        # --- Xóa file local sau khi upload ---
        os.remove(path_file)
        # --- Tạo URL tải file ---
        url = minio_client.presigned_get_object(
            MINIO_BUCKET,
            object_name,
            expires=timedelta(hours=1)
        )
        return {
            "status": "success",
            "url": url
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }


def WarehouseStatistical_Dashboard_function(conn, input):
    try:
        list_data = {}
        chart_data = {}
        cursor = conn.cursor()
        ### ---------------------------------------------------------------------------------------- 
        ### Point data ###
        ### ----------------------------------------------------------------------------------------
        # Tổng số lượng hàng hóa
        query_total_product = '''
            SELECT 
                COUNT(DISTINCT "id") AS total_product
            FROM "WS_Import"
        '''
        cursor.execute(query_total_product)
        total_product = cursor.fetchone()
        # Tổng số lượng nhập trong ngày
        query_import_by_date = '''
            SELECT
                COUNT(DISTINCT "id") AS import_by_date
            FROM "WS_Import"
            WHERE DATE("import_time") <= %s AND DATE("import_time") >= %s 
        '''
        cursor.execute(query_import_by_date, (input.filter.datetime_end, input.filter.datetime_start))
        import_by_date = cursor.fetchone()
        # Tổng số lượng xuất trong ngày
        query_export_by_date = '''
            SELECT
                COUNT(DISTINCT "id") AS export_by_date
            FROM "WS_Export"
            WHERE DATE("export_time") <= %s AND DATE("export_time") >= %s 
        '''
        cursor.execute(query_export_by_date, (input.filter.datetime_end, input.filter.datetime_start))
        export_by_date = cursor.fetchone()

        # Tổng số lượng chưa lắp đặt trong ngày
        query_not_installation_by_date = '''
            SELECT
                COUNT(DISTINCT "id") AS installation_by_date
            FROM "WS_Installation"
            WHERE "seri_number" IS NULL
        '''
        cursor.execute(query_not_installation_by_date)
        not_installation_by_date = cursor.fetchone()

        # Tổng số PO
        query_total_PO = '''
            SELECT
                COUNT(DISTINCT "import_id") AS total_PO,
                COUNT(DISTINCT "project_code") AS total_project
            FROM "WS_Import"
        '''
        cursor.execute(query_total_PO)
        row = cursor.fetchone()
        total_PO = row[0]
        total_project = row[1]

        point_data = {
            "total_product": total_product[0],
            "import_by_date": import_by_date[0],
            "export_by_date": export_by_date[0],
            "not_installation_by_date": not_installation_by_date[0],
            "total_PO": total_PO,
            "total_project": total_project
        }

        ### ---------------------------------------------------------------------------------------- 
        ### List data ###
        ### ----------------------------------------------------------------------------------------
        # List nhập hàng theo dự án
        query_list_import = '''
            SELECT
                "project_code",
                SUM("quantity") AS total_quantity
            FROM "WS_Import"
            WHERE DATE("import_time") BETWEEN %s AND %s
            GROUP BY "project_code"
            ORDER BY "project_code";
        '''
        cursor.execute(query_list_import, (input.filter.datetime_start, input.filter.datetime_end))
        rows = cursor.fetchall()
        list_data["import"] = [
            {"project_code": row[0], "total_quantity": row[1]} for row in rows
        ]
        # List xuất hàng theo dự án
        query_list_export = '''
            SELECT
                "project_code",
                SUM("quantity") AS total_quantity
            FROM "WS_Export"
            WHERE DATE("export_time") BETWEEN %s AND %s
            GROUP BY "project_code"
            ORDER BY "project_code";
        '''
        cursor.execute(query_list_export, (input.filter.datetime_start, input.filter.datetime_end))
        rows = cursor.fetchall()
        list_data["export"] = [
            {"project_code": row[0], "total_quantity": row[1]} for row in rows
        ]
        # List lắp đặt theo dự án
        query_list_installation = '''
            SELECT
                "project_code",
                COUNT(DISTINCT "id") AS total_quantity
            FROM "WS_Installation"
            WHERE "seri_number" IS NOT NULL
            GROUP BY "project_code"
            ORDER BY "project_code";
        '''
        cursor.execute(query_list_installation, (input.filter.datetime_start, input.filter.datetime_end))
        rows = cursor.fetchall()
        list_data["installation"] = [
            {"project_code": row[0], "total_quantity": row[1]} for row in rows
        ]
        ### ----------------------------------------------------------------------------------------
        ### Chart data ###
        ### ----------------------------------------------------------------------------------------
        # Chart tròn nhập xuất hàng theo ngày
        chart_data['pie_chart'] = {
            "import_quantity": import_by_date[0],
            "export_quantity": export_by_date[0]
        }

        list_range = ['day', 'week', 'month', 'quarter', 'year']
        format_map = {
            'day': 'YYYY-MM-DD',
            'week': 'IYYY-"W"IW',
            'month': 'YYYY-MM',
            'quarter': 'YYYY-"Q"Q',
            'year': 'YYYY'
        }
        chart_data['bar_chart'] = {}
        for range_type in list_range:
            date_format = format_map[range_type]
            query_import_export_by_date = f'''
                SELECT
                    TO_CHAR(DATE_TRUNC('{range_type}', t.time), '{date_format}') AS period,
                    SUM(CASE WHEN t.type = 'import' THEN t.quantity ELSE 0 END) AS total_import,
                    SUM(CASE WHEN t.type = 'export' THEN t.quantity ELSE 0 END) AS total_export
                FROM (
                    SELECT "import_time" AS time, "quantity", 'import' AS type FROM "WS_Import"
                    UNION ALL
                    SELECT "export_time" AS time, "quantity", 'export' AS type FROM "WS_Export"
                ) t
                GROUP BY TO_CHAR(DATE_TRUNC('{range_type}', t.time), '{date_format}')
                ORDER BY period;
            '''
            cursor.execute(query_import_export_by_date)
            rows = cursor.fetchall()
            chart_data['bar_chart'][range_type] = {
                "datetime_data": [row[0] for row in rows],
                "import_data": [row[1] for row in rows],
                "export_data": [row[2] for row in rows],
            }
        
        return {
            "status": "success",
            "point": point_data,
            "list": list_data,
            "chart": chart_data
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }
    finally:
        cursor.close()